import random
from openpyxl import load_workbook as load  # To read/modify Excel file
from urllib.request import urlopen as uReq
from bs4 import BeautifulSoup as soup  # BeautifulSoup ver 4
from itranslate import itranslate as itrans # Google translator API

# Open Source Excel file
srcFile = "Updated_List.xlsx"
srcWB = load(filename=srcFile)
srcWBSheet = srcWB['Sheet1']

# Open Excel file to be updated
file = "Data.xlsx"
wb = load(filename=file)
wbSheet = wb['Sheet1']

# Iterate every row from source file
for i, row in enumerate(srcWBSheet.iter_rows()):
    korName = row[0].internal_value  # 한글이름
    if i > 0:
        if korName is not None:
            print('Updating '+korName)
            found = False
            for wbRow in wbSheet.iter_rows():
                # Find if KorName exists in the file
                if wbRow[1].internal_value == korName:
                    found = True
                    wbRow[3].value = row[2].value #등장연도
                    wbRow[4].value = row[3].value  # 탄생연도
                    wbRow[5].value = row[4].value + random.randint(-3, 3)  # 통솔
                    if wbRow[5].value > 100:
                        wbRow[5].value = 100
                    if wbRow[5].value < 1:
                        wbRow[5].value = 1
                    wbRow[6].value = row[5].value + random.randint(-3, 3)  # 무력
                    if wbRow[6].value > 100:
                        wbRow[6].value = 100
                    if wbRow[6].value < 1:
                        wbRow[6].value = 1
                    wbRow[7].value = row[6].value + random.randint(-3, 3)  # 지력
                    if wbRow[7].value > 100:
                        wbRow[7].value = 100
                    if wbRow[7].value < 1:
                        wbRow[7].value = 1
                    wbRow[8].value = row[7].value + random.randint(-3, 3)  # 정치
                    if wbRow[8].value > 100:
                        wbRow[8].value = 100
                    if wbRow[8].value < 1:
                        wbRow[8].value = 1
                    wbRow[9].value = row[8].value + random.randint(-3, 3)  # 매력
                    if wbRow[9].value > 100:
                        wbRow[9].value = 100
                    if wbRow[9].value < 1:
                        wbRow[9].value = 1
                    wbRow[10].value = wbRow[6].value + wbRow[7].value + wbRow[9].value # 무력+매력+지력
                    wbRow[11].value = wbRow[10].value + wbRow[5].value + wbRow[8].value # 전체 능력 종합
                    break
            
            # If data not found. Randomly generate
            if found == False:
                print(korName + ' 추가중')
                newRow = wbSheet[wbSheet.max_row+1]
                newRow[0].value = row[12].value # 영어이름
                if newRow[0].value is None:
                    # Use google translator 
                    transName = itrans(row[1].value, to_lang="en")
                    newRow[0].value = transName
                newRow[1].value = row[0].value # 한글이름
                newRow[2].value = row[1].value  # 한문이름
                newRow[3].value = row[2].value  # 등장연도
                newRow[4].value = row[3].value  # 탄생연도
                newRow[5].value = row[4].value + random.randint(-3, 3)  # 통솔
                if newRow[5].value > 100:
                    newRow[5].value = 100
                if newRow[5].value < 1:
                    newRow[5].value = 1
                newRow[6].value = row[5].value + random.randint(-3, 3)  # 무력
                if newRow[6].value > 100:
                    newRow[6].value = 100
                if newRow[6].value < 1:
                    newRow[6].value = 1
                newRow[7].value = row[6].value + random.randint(-3, 3)  # 지력
                if newRow[7].value > 100:
                    newRow[7].value = 100
                if newRow[7].value < 1:
                    newRow[7].value = 1
                newRow[8].value = row[7].value + random.randint(-3, 3)  # 정치
                if newRow[8].value > 100:
                    newRow[8].value = 100
                if newRow[8].value < 1:
                    newRow[8].value = 1
                newRow[9].value = row[8].value + random.randint(-3, 3)  # 매력
                if newRow[9].value > 100:
                    newRow[9].value = 100
                if newRow[9].value < 1:
                    newRow[9].value = 1
                newRow[10].value = newRow[6].value + newRow[7].value + newRow[9].value  # 무력+매력+지력
                newRow[11].value = newRow[10].value + newRow[5].value + newRow[8].value  # 전체 능력 종합


# Iterate every row in the file to be updated and fill missing data randomly
for i, row in enumerate(wbSheet.iter_rows()):
    korName = row[0].internal_value  # 한글이름
    if i > 0:
        if row[3].value is None:
            print(korName + ' 랜덤 데이터 생성중')
            row[3].value = random.randint(187, 248)  # 등장연도
            row[4].value = random.randint(165, 210)  # 탄생연도
            row[5].value = random.randint(3, 80)  # 통솔
            row[6].value = random.randint(3, 80)  # 무력
            row[7].value = random.randint(3, 80)  # 지력
            row[8].value = random.randint(3, 80)  # 정치
            row[9].value = random.randint(3, 80)  # 매력
            row[10].value = row[6].value + row[7].value + row[9].value  # 무력+매력+지력
            row[11].value = row[10].value + row[5].value + row[8].value  # 전체 능력 종합

wb.save(file)
print('All done.')
