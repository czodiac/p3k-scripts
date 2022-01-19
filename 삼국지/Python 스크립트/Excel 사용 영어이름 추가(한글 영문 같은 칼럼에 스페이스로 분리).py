# https://the3kingdoms.fandom.com/wiki/Special:AllPages?from=Shishun+Rui

from openpyxl import load_workbook as load  # To read/modify Excel file
from urllib.request import urlopen as uReq
from bs4 import BeautifulSoup as soup  # BeautifulSoup ver 4

# Open Source Excel file
srcFile = "EngCharNames_30.xlsx"
srcWB = load(filename=srcFile)
srcWBSheet = srcWB['Sheet5']

# Open Excel file to be updated
file = "Updated_List.xlsx"
wb = load(filename=file)
wbSheet = wb['Sheet1']


for row in srcWBSheet.iter_rows():
    toFind = row[0].internal_value
    if toFind is not None:
        toFind = toFind.replace('\xa0', ' ')
        blankIndex = toFind.find(' ')
        korName = toFind[0:blankIndex]
        if toFind.find('(') > -1:
            engName = toFind[blankIndex+1:toFind.find('(')].strip()
        else:
            engName = toFind[blankIndex+1:].replace('\xa0', '').strip()

        # Find if KorName exists in the file
        for row in wbSheet.iter_rows():
            if row[0].internal_value == korName:
                if row[12].value is None:
                    row[12].value = engName
                    print(engName + ' updated')
                else:
                    if row[12].value != engName:
                        value = input(
                            "현재 이름:" + row[12].value + ", 새 이름: " + engName + ". Enter Y to update. N to skip: \n")
                        if value == 'Y':
                            row[12].value = engName
                            print(engName + ' updated')
                        else:
                            print(engName + ' NOT updated')

                wb.save('Updated_List.xlsx')
                break

print('All done.')
